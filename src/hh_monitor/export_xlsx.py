from __future__ import annotations

from collections.abc import Mapping
from pathlib import Path

from openpyxl import Workbook

from .models import ChangeRow, SeenVacancyRow

HEADERS = [
    "date_seen",
    "vacancy_id",
    "title",
    "company",
    "salary",
    "area",
    "url",
    "match_reason",
    "status",
]


def _write_sheet(ws: object, rows: list[ChangeRow]) -> None:
    ws.append(HEADERS)
    for row in rows:
        ws.append(
            [
                row.date_seen,
                row.vacancy_id,
                row.title,
                row.company,
                row.salary,
                row.area,
                row.url,
                row.match_reason,
                row.status,
            ]
        )


def _write_mapping_sheet(
    ws: object,
    headers: list[str],
    rows: list[Mapping[str, str]],
) -> None:
    ws.append(headers)
    for row in rows:
        ws.append([row.get(key, "") for key in headers])


def export_changes_xlsx(rows: list[ChangeRow], out_path: Path) -> Path:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()

    ws_all = workbook.active
    ws_all.title = "all"
    _write_sheet(ws_all, rows)

    for status in ["new", "updated", "removed"]:
        ws = workbook.create_sheet(title=status)
        _write_sheet(ws, [row for row in rows if row.status == status])

    workbook.save(out_path)
    return out_path


def export_ui_tables_xlsx(
    *,
    search_rows: list[SeenVacancyRow],
    deep_rows: list[Mapping[str, str]],
    out_path: Path,
) -> Path:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()

    search_headers = [
        "vacancy_id",
        "title",
        "company",
        "published_at",
        "area",
        "location_detail",
        "salary",
        "pre_dd_meta",
        "data_source_status",
        "url",
        "match_reason",
        "parse_status",
        "date_seen",
    ]
    deep_headers = [
        "vacancy_id",
        "title",
        "company",
        "url",
        "status",
        "added_at",
        "last_result_at",
    ]

    ws_search = workbook.active
    ws_search.title = "Общий поиск"
    search_payload = [
        {
            "vacancy_id": row.vacancy_id,
            "title": row.title,
            "company": row.company,
            "published_at": row.published_at,
            "area": row.area,
            "location_detail": row.location_detail,
            "salary": row.salary,
            "pre_dd_meta": row.pre_dd_meta,
            "data_source_status": row.data_source_status,
            "url": row.url,
            "match_reason": row.match_reason,
            "parse_status": row.parse_status,
            "date_seen": row.date_seen,
        }
        for row in search_rows
    ]
    _write_mapping_sheet(ws_search, search_headers, search_payload)

    ws_deep = workbook.create_sheet(title="Deep-dive")
    _write_mapping_sheet(ws_deep, deep_headers, deep_rows)

    workbook.save(out_path)
    return out_path
