from __future__ import annotations

from openpyxl import load_workbook

from hh_monitor.export_xlsx import export_ui_tables_xlsx
from hh_monitor.models import SeenVacancyRow


def test_export_ui_tables_xlsx_creates_search_and_deep_sheets(tmp_path) -> None:
    out_path = tmp_path / "report.xlsx"
    search_rows = [
        SeenVacancyRow(
            date_seen="2026-02-15T22:00:00+00:00",
            vacancy_id="123",
            title="Chief of Staff",
            company="ACME",
            salary="500000",
            area="Москва",
            url="https://hh.ru/vacancy/123",
            match_reason="include:chief of staff",
            parse_status="new",
            published_at="2026-02-15T18:00:00+03:00",
            location_detail="Москва, Пресненская набережная, 10",
            pre_dd_meta="Более 6 лет · Полная занятость",
            data_source_status="api_detail",
        )
    ]
    deep_rows = [
        {
            "vacancy_id": "123",
            "title": "Chief of Staff",
            "company": "ACME",
            "url": "https://hh.ru/vacancy/123",
            "status": "done",
            "added_at": "2026-02-15T22:00:00+00:00",
            "last_result_at": "2026-02-15T22:10:00+00:00",
        }
    ]

    export_ui_tables_xlsx(search_rows=search_rows, deep_rows=deep_rows, out_path=out_path)

    workbook = load_workbook(out_path)
    assert "Общий поиск" in workbook.sheetnames
    assert "Deep-dive" in workbook.sheetnames

    ws_search = workbook["Общий поиск"]
    assert ws_search.cell(1, 1).value == "vacancy_id"
    assert ws_search.cell(2, 1).value == "123"
    assert ws_search.cell(2, 2).value == "Chief of Staff"
    assert ws_search.cell(2, 4).value == "2026-02-15T18:00:00+03:00"
    assert ws_search.cell(2, 8).value == "Более 6 лет · Полная занятость"

    ws_deep = workbook["Deep-dive"]
    assert ws_deep.cell(1, 1).value == "vacancy_id"
    assert ws_deep.cell(2, 1).value == "123"
    assert ws_deep.cell(2, 5).value == "done"
